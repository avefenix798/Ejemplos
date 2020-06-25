

# importamos load_workbook
from openpyxl import load_workbook

# ruta de nuestro archivo
filesheet = "./demosheet.xlsx"

# creamos el objeto load_workbook
wb = load_workbook(filesheet)

# Seleccionamos el archivo
sheet = wb.active

# Ingresamos el valor 56 en la celda 'A1'
sheet['A1'] = 56
sheet['A2'] = 123

# Ingresamos el valor 1845 en la celda 'B3'
sheet['B3'] = 1845

# Guardamos el archivo con los cambios
wb.save(filesheet)

print("Termine de Escribir ");

